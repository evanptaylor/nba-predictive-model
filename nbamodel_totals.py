import requests
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from datetime import datetime, timedelta
from sklearn.model_selection import train_test_split, GridSearchCV
from sklearn.ensemble import RandomForestClassifier, VotingRegressor
from sklearn.preprocessing import StandardScaler
from sklearn.ensemble import RandomForestRegressor
from sklearn.linear_model import LogisticRegression
from sklearn.linear_model import LinearRegression
from xgboost import XGBRegressor
from sklearn.metrics import accuracy_score, mean_squared_error, r2_score
from nba_api.stats.static import teams
from nba_api.stats.endpoints import leaguegamelog, teamestimatedmetrics
import config

API_KEY = config.api_key

def fetch_odds_data(api_key):
    url = "https://api.the-odds-api.com/v4/sports/basketball_nba/odds/"

    # Add parameters to a dictionary
    params = {
        'api_key': api_key,
        'regions': 'us',
        'markets': 'totals',
        'oddsFormat': 'american',
        'dateFormat': 'iso',
        'bookmakers': ['bovada'],
        #'bookmakers': 1
    }

    response = requests.get(url, params=params)

    if response.status_code == 200:
        data = response.json()
    else:
        raise Exception(f"Failed to fetch odds data. Error: {response.text}")
    odds_list = []
    for game in data:
        if game is None:
            continue
        if not game['bookmakers']:
            continue

        bookmaker = game['bookmakers'][0]
        if not bookmaker['markets']:
            continue

        market = bookmaker['markets'][0]
        if not market['outcomes']:
            continue

        outcomes = market['outcomes']
        game_data = {
            
            'game_id': game['id'],
            'team1_name': game['home_team'],
            'team2_name': game['away_team'],
            'over_odds': outcomes[0]['price'],
            'under_odds': outcomes[1]['price'],
            'total_points': outcomes[0]['point']
        }
        odds_list.append(game_data)
    return pd.DataFrame(odds_list)

def fetch_game_data(start_date, end_date):
    nba_teams = teams.get_teams()

    game_log = leaguegamelog.LeagueGameLog(date_from_nullable=start_date, date_to_nullable=end_date, league_id='00')
    games = game_log.get_data_frames()[0]
    nba_games = games[games['TEAM_ID'].isin([team['id'] for team in nba_teams])]

    return pd.DataFrame(nba_games)

def fetch_team_stats_data(start_date, end_date):
    nba_teams = teams.get_teams()

    # Fetch NBA game logs for all time
    #game_log = leaguegamelog.LeagueGameLog(date_from_nullable=start_date, date_to_nullable=end_date, league_id='00')
    game_log = leaguegamelog.LeagueGameLog(season='2022-23', league_id='00')
    game_log_e = teamestimatedmetrics.TeamEstimatedMetrics(season='2022-23', league_id='00')
    games = game_log.get_data_frames()[0]
    games_e = game_log_e.get_data_frames()[0]
    nba_games = games[games['TEAM_ID'].isin([team['id'] for team in nba_teams])]
    nba_games_e = games_e[games_e['TEAM_ID'].isin([team['id'] for team in nba_teams])]
    #games_21_22 = games[games.SEASON_ID.str[-4:] == '2021']

    #print(nba_games)
    # Calculate the stats for each team
    team_stats = []
    for team in nba_teams:
        team_id = team['id']
        team_games = nba_games[(nba_games['TEAM_ID'] == team_id)]
        team_games_e = nba_games_e[(nba_games_e['TEAM_ID'] == team_id)]
        #wonlast = 1 if team_games[team_games['WL'][len(team_games[team_games['WL']])-1] == 'W'] else 0
        wins = len(team_games[team_games['WL'] == 'W'])
        losses = len(team_games[team_games['WL'] == 'L'])
        win_percentage = wins / (wins + losses) if (wins + losses) > 0 else 0

        points = team_games['PTS'].sum()
        games_played = len(team_games)
        average_points = points / games_played if games_played > 0 else 0
        fg3_pct = team_games['FG3_PCT'].sum() / games_played if games_played > 0 else 0
        fg2_pct = team_games['FG_PCT'].sum() / games_played if games_played > 0 else 0
        avg_reb = team_games['REB'].sum() / games_played if games_played > 0 else 0
        avg_oreb = team_games['OREB'].sum() / games_played if games_played > 0 else 0
        avg_ast = team_games['AST'].sum() / games_played if games_played > 0 else 0
        avg_tov =team_games['TOV'].sum() / games_played if games_played > 0 else 0
        plus_minus = team_games['PLUS_MINUS'].sum() / games_played if games_played > 0 else 0
        fgm = team_games['FGM'].sum() / games_played if games_played > 0 else 0
        fg3m = team_games['FG3M'].sum() / games_played if games_played > 0 else 0
        fga = team_games['FGA'].sum() / games_played if games_played > 0 else 0
        fta = team_games['FTA'].sum() / games_played if games_played > 0 else 0
        pace = team_games_e["E_PACE"].sum()
        off_rating = team_games_e["E_OFF_RATING"].sum()
        efg_pct = (fgm + 0.5*fg3m) / fga
        ftr = fta/fga
        net_rating = team_games_e["E_NET_RATING"].sum()
        ast_ratio = team_games_e["E_AST_RATIO"].sum()
        tov_pct = team_games_e["E_TM_TOV_PCT"].sum()
        #eFG% = (FGM + 0.5 * FG3M) / FGA
        team_stat = {
            'TEAM_ID': team_id,
            'team_name': team['full_name'],
            'average_points': average_points,
            'fg3_pct': fg3_pct,
            'oreb':avg_oreb,
            'ast':avg_ast,
            'tov':avg_tov,
            'plus_minus': plus_minus,
            'efg_pct': efg_pct,
            'ft_rate': ftr,
            'pace': pace,
            'off_rating': off_rating,
            'net_rating': net_rating,
            'ast_ratio': ast_ratio,
            'tov_pct':tov_pct
        }
        team_stats.append(team_stat)
    return pd.DataFrame(team_stats)

def preprocess_data(team_stats_data, odds_data):

    for team in odds_data['team1_name']:
        merged_df = odds_data.merge(team_stats_data, left_on="team1_name", right_on="team_name", suffixes=('', '_team1'))
        merged_df = merged_df.merge(team_stats_data, left_on="team2_name", right_on="team_name", suffixes=('_team1', '_team2'))
        merged_df = merged_df.drop(columns=['TEAM_ID_team1', 'team_name_team1', 'TEAM_ID_team2', 'team_name_team2', 'game_id', 'team1_name','team2_name'])

    return merged_df

    
def make_predictions(model, X_test):
    predictions = model.predict(X_test)
    return predictions

def predict_upcoming_games(model, upcoming_odds_data, team_data):
    upcoming_data = preprocess_data(team_data, odds_data)

    # Remove the 'total_points' column from the upcoming_data
    X_upcoming = upcoming_data.drop(columns=['total_points'])

    predictions = make_predictions(model, X_upcoming)
    
    # Combine the game information with predictions
    upcoming_game_predictions = upcoming_odds_data[['team1_name', 'team2_name', 'over_odds', 'under_odds', 'total_points']].copy()
    upcoming_game_predictions['predicted_total_points'] = predictions
    
    return upcoming_game_predictions



def train_stacked_model(X_train, y_train):
    rf = RandomForestRegressor()
    lr = LogisticRegression(max_iter=3000)
    xgb = XGBRegressor()

    model = VotingRegressor(
        estimators=[('rf', rf), ('lr', lr), ('xgb', xgb)],
        weights=[1, 1, 1]
    )
    model.fit(X_train, y_train)
    return model

def train_model(X_train, y_train):
    # Perform hyperparameter tuning
    param_grid = {
        'n_estimators': [100, 200, 300],
        'max_depth': [None, 10, 20, 30],
        'min_samples_split': [2, 5, 10],
        'min_samples_leaf': [1, 2, 4]
    }
    model = RandomForestRegressor()
    grid_search = GridSearchCV(model, param_grid, cv=5, scoring='neg_mean_squared_error', n_jobs=-1)
    grid_search.fit(X_train, y_train)
    return grid_search.best_estimator_


def format_excel(file):
    # Load the workbook and select the sheet
    wb = load_workbook(file)
    ws = wb.active

    # Define the fill colors for the winner and loser
    winner_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    loser_fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB", fill_type="solid")

    # Iterate through the rows and apply the fill colors
    for row in range(2, ws.max_row + 1):
        line = ws.cell(row=row, column=5).value
        model = ws.cell(row=row, column=6).value
        
        if model > line:
            ws.cell(row=row, column=3).fill = winner_fill
            ws.cell(row=row, column=4).fill = loser_fill
        else:
            ws.cell(row=row, column=3).fill = loser_fill
            ws.cell(row=row, column=4).fill = winner_fill
        
    for col in range(1, 26):
        ws.column_dimensions[chr(col + 64)].width = 20
    # Save the formatted workbook
    wb.save("predictions_totals.xlsx")

if __name__ == "__main__":
    start_date = "2022-10-19"
    end_date = "2023-03-30"
    odds_data = fetch_odds_data(API_KEY)
    team_data = fetch_team_stats_data(start_date, end_date)
    data = preprocess_data(team_data, odds_data)
    #print(odds_data)
    #print(team_data)
    X = data.drop(columns=['total_points'])
    y = data['total_points']
    
    X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, random_state=42)

    # Train the model
    model = train_model(X_train, y_train)

    # Make predictions
    predictions = make_predictions(model, X_test)

    # Calculate Mean Squared Error and R-squared score
    mse = mean_squared_error(y_test, predictions)
    r2 = r2_score(y_test, predictions)
    print(f"Model Mean Squared Error: {mse}")
    print(f"Model R-squared score: {r2}")

    # Fetch upcoming games odds data (change the start_date and end_date to the desired range)
    upcoming_odds_data = fetch_odds_data(API_KEY)

    # Predict upcoming games
    upcoming_game_predictions = predict_upcoming_games(model, upcoming_odds_data, team_data)
    print(upcoming_game_predictions)

    upcoming_game_predictions.to_excel("predictions_totals.xlsx", index=False)
    format_excel("predictions_totals.xlsx")
    print("Done! Open predictions_totals.xlsx to view")