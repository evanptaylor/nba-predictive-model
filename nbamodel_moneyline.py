import requests
import numpy as np
import pandas as pd
from nba_api.stats.static import teams
from nba_api.stats.endpoints import leaguegamelog, teamestimatedmetrics
from sklearn.model_selection import train_test_split, GridSearchCV
from sklearn.ensemble import RandomForestClassifier, VotingClassifier
from sklearn.linear_model import LogisticRegression
from sklearn.metrics import accuracy_score, confusion_matrix
from xgboost import XGBClassifier
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import config

#import api key from seperate file
API_KEY = config.api_key

#fetch moneyline odds for upcoming NBA games 
def fetch_odds_data(api_key):
    #the odds api 
    url = "https://api.the-odds-api.com/v4/sports/basketball_nba/odds/"
    
    #add odds and teams for upcoming games to a dataframe
    params = {
        'api_key': api_key,
        'regions': 'us',
        'markets': 'h2h',
        'oddsFormat': 'american',
        'dateFormat': 'iso',
        'bookmakers': ['draftkings']
    }
    response = requests.get(url, params=params)
    if response.status_code == 200:
        data = response.json()
    else:
        raise Exception(f"Failed to fetch odds data. Error: {response.text}")
    
    odds_list = []
    for game in data:
        game_data = {
            'game_id': game['id'],
            'team1_name': game['bookmakers'][0]['markets'][0]['outcomes'][0]['name'],
            'team2_name': game['bookmakers'][0]['markets'][0]['outcomes'][1]['name'],
            'team1_odds': game['bookmakers'][0]['markets'][0]['outcomes'][0]['price'],
            'team2_odds': game['bookmakers'][0]['markets'][0]['outcomes'][1]['price']
        }
        odds_list.append(game_data)
    return pd.DataFrame(odds_list)

#fetch historical team statisitcal data from nba_api library
def fetch_team_stats_data(start_date, end_date):
    #get game logs for all teams in date range
    nba_teams = teams.get_teams()
    game_log = leaguegamelog.LeagueGameLog(date_from_nullable=start_date, date_to_nullable=end_date, league_id='00')
    games = game_log.get_data_frames()[0]
    nba_games = games[games['TEAM_ID'].isin([team['id'] for team in nba_teams])]
    game_log_e = teamestimatedmetrics.TeamEstimatedMetrics(season='2022-23', league_id='00')
    games_e = game_log_e.get_data_frames()[0]
    nba_games_e = games_e[games_e['TEAM_ID'].isin([team['id'] for team in nba_teams])]

    #calculate releveant stats
    team_stats = []
    for team in nba_teams:
        team_id = team['id']
        team_games = nba_games[(nba_games['TEAM_ID'] == team_id)]
        team_games_e = nba_games_e[(nba_games_e['TEAM_ID'] == team_id)]
        wins = len(team_games[team_games['WL'] == 'W'])
        losses = len(team_games[team_games['WL'] == 'L'])
        win_percentage = wins / (wins + losses) if (wins + losses) > 0 else 0
        points = team_games['PTS'].sum()
        games_played = len(team_games)
        average_points = points / games_played if games_played > 0 else 0
        #fg3_pct = team_games['FG3_PCT'].sum() / games_played if games_played > 0 else 0
        #fg2_pct = team_games['FG_PCT'].sum() / games_played if games_played > 0 else 0
        avg_reb = team_games['REB'].sum() / games_played if games_played > 0 else 0
        avg_oreb = team_games['OREB'].sum() / games_played if games_played > 0 else 0
        avg_ast = team_games['AST'].sum() / games_played if games_played > 0 else 0
        avg_tov =team_games['TOV'].sum() / games_played if games_played > 0 else 0
        fgm = team_games['FGM'].sum() / games_played if games_played > 0 else 0
        fg3m = team_games['FG3M'].sum() / games_played if games_played > 0 else 0
        fga = team_games['FGA'].sum() / games_played if games_played > 0 else 0
        efg_pct = (fgm + 0.5*fg3m) / fga
        plus_minus = team_games['PLUS_MINUS'].sum() / games_played if games_played > 0 else 0
        pace = team_games_e["E_PACE"].sum()
        off_rating = team_games_e["E_OFF_RATING"].sum()
        net_rating = team_games_e["E_NET_RATING"].sum()
        ast_ratio = team_games_e["E_AST_RATIO"].sum()
        tov_pct = team_games_e["E_TM_TOV_PCT"].sum()

        #add relevant stats to dict and add a row for each team in a dataframe
        team_stat = {
            'TEAM_ID': team_id,
            'team_name': team['full_name'],
            'games_played': games_played,
            'wins': wins,
            'losses': losses,
            'win_percentage': win_percentage,
            'average_points': average_points,
            'reb':avg_reb,
            'oreb':avg_oreb,
            'ast':avg_ast,
            'tov':avg_tov,
            'plus_minus': plus_minus,
            'efg_pct': efg_pct,
            'pace': pace,
            'off_rating': off_rating, 
            'net_rating': net_rating,
            'ast_ratio': ast_ratio,
            'tov_pct':tov_pct
        }
        team_stats.append(team_stat)
    return pd.DataFrame(team_stats)

def preprocess_data(odds_data, team_stats_data):
    #merge odds and stats data for each team playing in upcoming games
    for team in odds_data['team1_name']:
        merged_df = odds_data.merge(team_stats_data, left_on="team1_name", right_on="team_name", suffixes=('', '_team1'))
        merged_df = merged_df.merge(team_stats_data, left_on="team2_name", right_on="team_name", suffixes=('_team1', '_team2'))
        merged_df = merged_df.drop(columns=['TEAM_ID_team1', 'team_name_team1', 'TEAM_ID_team2', 'team_name_team2'])
    return merged_df

#make data 'friendly' for the matrix operations
def select_features(data):
    #drop non-numerical columns for input matrix
    features_to_drop = ['game_id', 'team1_name', 'team2_name','team1_odds','team2_odds']
    return data.drop(columns=features_to_drop)

#employ a voting classifier to choose between regression models
def train_stacked_model(X_train, y_train):
    #train multiple models
    rf_params = {
        'n_estimators': [50, 100, 200],
        'max_depth': [None, 10, 20],
        'min_samples_split': [2, 5, 10],
        'min_samples_leaf': [1, 2, 4]
    }

    rf = RandomForestClassifier()
    grid_search = GridSearchCV(rf, rf_params, cv=2, n_jobs=-1)
    grid_search.fit(X_train, y_train)
    best_rf = grid_search.best_estimator_

    lr = LogisticRegression(max_iter=3000)
    xgb = XGBClassifier()

    #combine the models using a voting classifier
    model = VotingClassifier(
        estimators=[('rf', best_rf), ('lr', lr), ('xgb', xgb)],
        voting='soft'
    )
    model.fit(X_train, y_train)
    return model

def make_predictions(model, X_test):
    predictions = model.predict(X_test)
    probabilities = model.predict_proba(X_test)
    return predictions, probabilities

def predict_upcoming_games(model, upcoming_odds_data, team_stats_data):
    #preproccess upcoming game data
    upcoming_data = preprocess_data(upcoming_odds_data, team_stats_data)
    X_upcoming = select_features(upcoming_data)

    #make predictions and add to a dataframe
    predictions, probabilities = make_predictions(model, X_upcoming)
    upcoming_data['predicted_winner'] = np.where(predictions == 1, upcoming_data['team1_name'], upcoming_data['team2_name'])
    upcoming_data['team1_probability'] = probabilities[:, 1]
    upcoming_data['team2_probability'] = probabilities[:, 0]

    return upcoming_data

#output formatting
def format_excel(file):
    #load and format data into excel file
    wb = load_workbook(file)
    ws = wb.active
    winner_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    loser_fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB", fill_type="solid")

    for row in range(2, ws.max_row + 1):
        team1_prob = ws.cell(row=row, column=6).value
        team2_prob = ws.cell(row=row, column=7).value
        take_bet = ws.cell(row=row, column=8).value
        if team1_prob > team2_prob:
            ws.cell(row=row, column=1).fill = winner_fill
            ws.cell(row=row, column=2).fill = loser_fill
            ws.cell(row=row, column=3).fill = winner_fill
            ws.cell(row=row, column=4).fill = loser_fill
            ws.cell(row=row, column=5).fill = winner_fill
            ws.cell(row=row, column=6).fill = winner_fill
            ws.cell(row=row, column=7).fill = loser_fill
        else:
            ws.cell(row=row, column=1).fill = loser_fill
            ws.cell(row=row, column=2).fill = winner_fill
            ws.cell(row=row, column=3).fill = loser_fill
            ws.cell(row=row, column=4).fill = winner_fill
            ws.cell(row=row, column=5).fill = winner_fill
            ws.cell(row=row, column=6).fill = loser_fill
            ws.cell(row=row, column=7).fill = winner_fill
        if take_bet == 1:
            ws.cell(row=row, column=8).fill = winner_fill
        else:
            ws.cell(row=row, column=8).fill = loser_fill

    for col in range(1, 26):
        ws.column_dimensions[chr(col + 64)].width = 20

    wb.save("predictions.xlsx")

if __name__ == "__main__":
    #start and end date for training data
    #start_date = "2022-10-19" #start of nba season 2022-23
    start_date = "2023-02-19" #all star game 2023
    end_date = "2023-04-03"

    #fetch all data and preproccess
    odds_data = fetch_odds_data(API_KEY)
    team_stats_data = fetch_team_stats_data(start_date, end_date)
    team_ids = team_stats_data['TEAM_ID'].unique()
    data = preprocess_data(odds_data, team_stats_data)

    #set matrix and vector
    X = select_features(data)
    y = np.where(data['team1_odds'] < data['team2_odds'], 1, 0)

    #training and testing sets
    X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, random_state=42)

    #train model and calculate stats on the model
    model = train_stacked_model(X_train, y_train)
    predictions, probabilities = make_predictions(model, X_test)
    accuracy = accuracy_score(y_test, predictions)
    print(f"Model accuracy: {accuracy}")
    cm = confusion_matrix(y_test, predictions)
    print("Confusion matrix:\n", cm)

    #predict upcoming games
    upcoming_odds_data = fetch_odds_data(API_KEY)
    upcoming_game_predictions = predict_upcoming_games(model, upcoming_odds_data, team_stats_data)
    keep = ['team1_name','team2_name','team1_odds','team2_odds','predicted_winner','team1_probability','team2_probability']
    upcoming_game_predictions = upcoming_game_predictions.filter(keep)
    upcoming_game_predictions['certified_locks'] = np.where(
        ((upcoming_game_predictions['team1_probability'] > 0.75) & (upcoming_game_predictions['team1_odds'] > -250)) | 
        ((upcoming_game_predictions['team2_probability'] > 0.75) & (upcoming_game_predictions['team2_odds'] > -250)), 1, 0)
    upcoming_game_predictions.to_excel("predictions.xlsx", index=False)
    format_excel("predictions.xlsx")
    print("Done! Open predictions.xlsx to view")
    

